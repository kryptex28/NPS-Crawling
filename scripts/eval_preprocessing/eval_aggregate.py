import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

EVAL_DIR = Path(__file__).resolve().parents[2] / "evaluation" / "preprocessing"
RESULTS_DIR = EVAL_DIR / "results"
LABELING_XLSX = EVAL_DIR / "labeling.xlsx"

# Filename pattern: scores_<model_slug>__<ref_id>_metrics.csv
# The model slug itself can contain single underscores (e.g. "BAAI_bge-small-en-v1.5"),
# but the ``__`` (double underscore) before the ref_id is the unambiguous separator.
FILENAME_RE = re.compile(r"^scores_(?P<model>.+?)__(?P<ref>[^.]+)_metrics\.csv$")


def format_ref_id(ref_id: str) -> str:
    """Turn 'V2_short_definition' into 'V2: Short definition' for display."""
    if "_" not in ref_id:
        return ref_id
    prefix, _, rest = ref_id.partition("_")
    rest = rest.replace("_", " ")
    if rest:
        rest = rest[0].upper() + rest[1:]
    return f"{prefix}: {rest}"


def short_model_name(model_slug: str) -> str:
    """Strip vendor prefix and common 'sentence-transformers_' clutter."""
    name = model_slug
    for prefix in ("sentence-transformers_", "BAAI_"):
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    return name


def parse_metrics_csv(path: Path) -> dict | None:
    """Return the row with max F1 from one metrics CSV, plus timing."""
    best = None
    ms_per_ctx = None
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                f1 = float(row["F1"])
            except (KeyError, TypeError, ValueError):
                continue
            if ms_per_ctx is None:
                raw = row.get("embedding_ms_per_context", "")
                if raw not in ("", None):
                    try:
                        ms_per_ctx = float(raw)
                    except ValueError:
                        pass
            if best is None or f1 > best["F1"]:
                best = {
                    "threshold": float(row["threshold"]),
                    "TP": int(row["TP"]),
                    "FP": int(row["FP"]),
                    "FN": int(row["FN"]),
                    "TN": int(row["TN"]),
                    "precision": float(row["precision"]),
                    "recall": float(row["recall"]),
                    "F1": f1,
                }
    if best is None:
        return None
    best["embedding_ms_per_context"] = ms_per_ctx
    return best


def read_sweep(path: Path) -> list[dict]:
    """Read every threshold row from one metrics CSV, sorted by threshold."""
    rows = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                rows.append({
                    "threshold": float(row["threshold"]),
                    "precision": float(row["precision"]),
                    "recall": float(row["recall"]),
                    "F1": float(row["F1"]),
                    "TP": int(row["TP"]),
                    "FP": int(row["FP"]),
                })
            except (KeyError, TypeError, ValueError):
                continue
    rows.sort(key=lambda r: r["threshold"])
    return rows


def collect() -> list[dict]:
    """Walk every metrics CSV in EVAL_DIR and return one record per file."""
    records = []
    for path in sorted(EVAL_DIR.glob("scores_*_metrics.csv")):
        m = FILENAME_RE.match(path.name)
        if not m:
            # Legacy files (e.g. scores_<model>_metrics.csv without a __ref) skipped.
            print(f"  skipping (no reference text in name): {path.name}", file=sys.stderr)
            continue
        best = parse_metrics_csv(path)
        if best is None:
            print(f"  skipping (no rows): {path.name}", file=sys.stderr)
            continue
        records.append({
            "model_slug": m.group("model"),
            "model": short_model_name(m.group("model")),
            "reference_text_id": m.group("ref"),
            **best,
            "source_file": path.name,
        })
    return records


def write_long_csv(records: list[dict], path: Path) -> None:
    cols = [
        "model", "reference_text_id", "F1", "threshold",
        "precision", "recall", "TP", "FP", "FN", "TN",
        "embedding_ms_per_context", "model_slug", "source_file",
    ]
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        for r in records:
            writer.writerow({k: r.get(k, "") for k in cols})


def build_pivot(records: list[dict], value_fn) -> tuple[list[str], list[str], dict, dict]:
    """Pivot records into rows=model, cols=ref_id. Returns (row_labels, cols, cells, ms_per_model).

    ``value_fn(record)`` produces the cell value (string).
    """
    by_model: dict[str, dict[str, dict]] = defaultdict(dict)
    ms_per_model: dict[str, float | None] = {}
    for r in records:
        by_model[r["model"]][r["reference_text_id"]] = r
        # All ref texts share the same model timing — overwriting is fine.
        ms_per_model[r["model"]] = r.get("embedding_ms_per_context")

    # Preserve insertion order from the records list.
    models = list(by_model.keys())
    cols: list[str] = []
    seen = set()
    for r in records:
        rid = r["reference_text_id"]
        if rid not in seen:
            seen.add(rid)
            cols.append(rid)

    cells = {
        (model, ref): value_fn(by_model[model][ref])
        for model in models
        for ref in cols
        if ref in by_model[model]
    }
    return models, cols, cells, ms_per_model


def row_label(model: str, ms: float | None) -> str:
    if ms is None:
        return model
    return f"{model} ({ms:.1f} ms/ctx)"


def write_pivot_csv(records: list[dict], path: Path, value_fn) -> None:
    models, cols, cells, ms_per_model = build_pivot(records, value_fn)
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["model"] + cols)
        for m in models:
            writer.writerow(
                [row_label(m, ms_per_model[m])]
                + [cells.get((m, c), "") for c in cols]
            )


def print_markdown_table(records: list[dict], value_fn, title: str) -> None:
    models, cols, cells, ms_per_model = build_pivot(records, value_fn)
    if not models or not cols:
        print(f"\n{title}: no data")
        return
    headers = ["Model (embedding time)"] + cols
    rows = [
        [row_label(m, ms_per_model[m])] + [cells.get((m, c), "—") for c in cols]
        for m in models
    ]
    widths = [
        max(len(str(headers[i])), *(len(str(r[i])) for r in rows))
        for i in range(len(headers))
    ]
    def fmt_row(r):
        return "| " + " | ".join(str(r[i]).ljust(widths[i]) for i in range(len(r))) + " |"
    sep = "|" + "|".join("-" * (w + 2) for w in widths) + "|"
    print(f"\n### {title}\n")
    print(fmt_row(headers))
    print(sep)
    for r in rows:
        print(fmt_row(r))


def write_heatmap(records: list[dict], path: Path) -> None:
    """Render the F1 pivot as a heatmap with cell annotations."""
    try:
        import matplotlib
        matplotlib.use("Agg")  # headless backend; no Tk required
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError as exc:
        print(f"  Skipping heatmap ({exc}). Install matplotlib to enable.")
        return

    models, cols, cells, ms_per_model = build_pivot(
        records, value_fn=lambda r: r["F1"],
    )
    if not models or not cols:
        return

    matrix = np.array([
        [cells.get((m, c), np.nan) for c in cols]
        for m in models
    ], dtype=float)

    row_labels = [row_label(m, ms_per_model[m]) for m in models]

    fig_w = max(6.0, 1.4 * len(cols) + 4.0)
    fig_h = max(3.0, 0.7 * len(models) + 1.5)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))

    im = ax.imshow(matrix, aspect="auto", cmap="RdYlGn", vmin=0.0, vmax=1.0)

    ax.set_xticks(range(len(cols)))
    ax.set_xticklabels([format_ref_id(c) for c in cols], rotation=30, ha="right")
    ax.set_yticks(range(len(models)))
    ax.set_yticklabels(row_labels)

    # Annotate each cell with its F1 value (black on the red→green ramp stays readable).
    for i in range(len(models)):
        for j in range(len(cols)):
            val = matrix[i, j]
            if np.isnan(val):
                ax.text(j, i, "—", ha="center", va="center", color="black", fontsize=9)
                continue
            ax.text(j, i, f"{val:.3f}", ha="center", va="center",
                    color="black", fontsize=9)

    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)


def load_labels() -> dict[str, int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  Skipping separation histogram (openpyxl not installed).")
        return {}
    if not LABELING_XLSX.exists():
        print(f"  Skipping separation histogram (missing {LABELING_XLSX.name}).")
        return {}
    wb = load_workbook(LABELING_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    cid_col = headers.index("context_id")
    label_col = headers.index("label")
    labels: dict[str, int] = {}
    for r in rows:
        cid = r[cid_col]
        lbl = r[label_col]
        if cid is None or lbl is None or str(lbl).strip() == "":
            continue
        try:
            labels[str(cid)] = int(lbl)
        except (TypeError, ValueError):
            continue
    return labels


def load_scores_jsonl(path: Path) -> dict[str, float]:
    out: dict[str, float] = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            out[obj["context_id"]] = float(obj["similarity_score"])
    return out


def pick_top_models(records: list[dict], n: int = 2) -> list[dict]:
    """For each model, find its best (ref_text, F1, threshold). Return top-n by F1."""
    best_per_model: dict[str, dict] = {}
    for r in records:
        cur = best_per_model.get(r["model"])
        if cur is None or r["F1"] > cur["F1"]:
            best_per_model[r["model"]] = r
    return sorted(best_per_model.values(), key=lambda x: x["F1"], reverse=True)[:n]


def write_separation_histograms(records: list[dict], path: Path) -> None:
    """Plot positive-vs-negative score distributions for the top-N models."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError as exc:
        print(f"  Skipping separation histogram ({exc}).")
        return

    labels = load_labels()
    if not labels:
        return

    top = pick_top_models(records, n=2)
    if not top:
        return

    fig, axes = plt.subplots(1, len(top), figsize=(6.0 * len(top), 4.2), sharey=True)
    if len(top) == 1:
        axes = [axes]

    bins = np.linspace(0.0, 1.0, 41)

    for ax, rec in zip(axes, top):
        scores_file = EVAL_DIR / f"scores_{rec['model_slug']}__{rec['reference_text_id']}.jsonl"
        if not scores_file.exists():
            ax.set_title(f"{rec['model']} (scores file missing)")
            continue
        scores = load_scores_jsonl(scores_file)
        pos = [scores[cid] for cid, lbl in labels.items() if lbl == 1 and cid in scores]
        neg = [scores[cid] for cid, lbl in labels.items() if lbl == 0 and cid in scores]

        ax.hist(neg, bins=bins, alpha=0.55, color="#d62728",
                label=f"Negative (n={len(neg)})", edgecolor="white", linewidth=0.3)
        ax.hist(pos, bins=bins, alpha=0.55, color="#2ca02c",
                label=f"Positive (n={len(pos)})", edgecolor="white", linewidth=0.3)
        ax.axvline(rec["threshold"], color="black", linestyle="--", linewidth=1.2,
                   label=f"Best threshold = {rec['threshold']:.2f}")
        ax.set_xlabel("Cosine similarity")
        ax.set_xlim(0.0, 1.0)
        ax.set_title(
            f"{rec['model']}\n{format_ref_id(rec['reference_text_id'])}  |  "
            f"F1 = {rec['F1']:.3f}  (P={rec['precision']:.2f}, R={rec['recall']:.2f})"
        )
        ax.legend(loc="upper left", fontsize=8)
        ax.grid(axis="y", alpha=0.25)

    axes[0].set_ylabel("Number of context windows")
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def write_threshold_sweep(records: list[dict], path: Path, n: int = 2) -> None:
    """Plot precision / recall / F1 against threshold for the top-N (model, ref) combos.

    Justifies the chosen operating point: shows whether the F1 peak is a sharp
    spike or a broad plateau, and how precision trades against recall as the
    threshold moves. Reads the full sweep straight from each combo's metrics CSV.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError as exc:
        print(f"  Skipping threshold sweep ({exc}).")
        return

    top = pick_top_models(records, n=n)
    if not top:
        return

    fig, axes = plt.subplots(1, len(top), figsize=(6.0 * len(top), 4.2), sharey=True)
    if len(top) == 1:
        axes = [axes]

    for ax, rec in zip(axes, top):
        sweep = read_sweep(EVAL_DIR / rec["source_file"])
        if not sweep:
            ax.set_title(f"{rec['model']} (metrics CSV missing)")
            continue
        thr = [r["threshold"] for r in sweep]
        # Precision is 0/0 (undefined) once nothing is predicted positive; the CSV
        # records it as 0.0. Mask those points so the line ends instead of diving
        # to zero, which would misread as a precision collapse.
        precision = [
            r["precision"] if (r["TP"] + r["FP"]) > 0 else np.nan
            for r in sweep
        ]
        ax.plot(thr, precision, color="#1f77b4", label="Precision")
        ax.plot(thr, [r["recall"] for r in sweep], color="#ff7f0e", label="Recall")
        ax.plot(thr, [r["F1"] for r in sweep], color="#2ca02c", linewidth=2.2, label="F1")
        ax.axvline(rec["threshold"], color="black", linestyle="--", linewidth=1.2,
                   label=f"Best F1 @ {rec['threshold']:.2f}")
        ax.set_xlabel("Threshold")
        ax.set_xlim(min(thr), max(thr))
        ax.set_ylim(0.0, 1.02)
        ax.set_title(
            f"{rec['model']}\n{format_ref_id(rec['reference_text_id'])}  |  "
            f"best F1 = {rec['F1']:.3f}"
        )
        ax.legend(loc="lower center", fontsize=8)
        ax.grid(alpha=0.25)

    axes[0].set_ylabel("Score")
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def write_speed_quality_scatter(records: list[dict], path: Path) -> None:
    """Scatter best-F1 against embedding cost (ms/context), one point per model.

    Makes the model-choice trade-off visual: a model that is both cheap and
    top-scoring dominates the larger, slower ones, so paying for a bigger model
    is only worth it if it buys F1. Uses each model's best reference text.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError as exc:
        print(f"  Skipping speed/quality scatter ({exc}).")
        return

    # Best (max-F1) combo per model.
    best_per_model: dict[str, dict] = {}
    for r in records:
        cur = best_per_model.get(r["model"])
        if cur is None or r["F1"] > cur["F1"]:
            best_per_model[r["model"]] = r
    points = [r for r in best_per_model.values() if r.get("embedding_ms_per_context")]
    if not points:
        print("  Skipping speed/quality scatter (no timing data).")
        return

    fig, ax = plt.subplots(figsize=(7.5, 5.0))
    xs = [r["embedding_ms_per_context"] for r in points]
    ys = [r["F1"] for r in points]
    ax.scatter(xs, ys, s=80, color="#1f77b4", zorder=3)

    for r in points:
        ax.annotate(
            f"{r['model']}\n({format_ref_id(r['reference_text_id'])})",
            (r["embedding_ms_per_context"], r["F1"]),
            textcoords="offset points", xytext=(9, 4), fontsize=8,
        )

    ax.set_xscale("log")
    # Headroom so the point annotations don't collide with the frame.
    ax.set_xlim(min(xs) * 0.7, max(xs) * 2.2)
    y_lo, y_hi = min(ys), max(ys)
    pad = (y_hi - y_lo) * 0.15 or 0.02
    ax.set_ylim(y_lo - pad, y_hi + pad)
    ax.set_xlabel("Embedding cost  (ms / context, log scale)")
    ax.set_ylabel("Best F1")
    ax.set_title("Embedding model: quality vs. cost  (best reference text per model)")
    ax.grid(alpha=0.25, which="both")
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def main():
    records = collect()
    if not records:
        print("No metrics CSVs found. Run eval_evaluate.py first.")
        return

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    long_path = RESULTS_DIR / "summary_long.csv"
    f1_path = RESULTS_DIR / "summary_best_f1.csv"
    f1_thr_path = RESULTS_DIR / "summary_best_f1_at_threshold.csv"
    heatmap_path = RESULTS_DIR / "heatmap_best_f1.png"
    hist_path = RESULTS_DIR / "score_separation_top_models.png"
    sweep_path = RESULTS_DIR / "threshold_sweep_top_models.png"
    scatter_path = RESULTS_DIR / "speed_vs_quality.png"

    write_long_csv(records, long_path)
    write_pivot_csv(records, f1_path, value_fn=lambda r: f"{r['F1']:.3f}")
    write_pivot_csv(
        records, f1_thr_path,
        value_fn=lambda r: f"{r['F1']:.3f} @ {r['threshold']:.2f}",
    )
    write_heatmap(records, heatmap_path)
    write_separation_histograms(records, hist_path)
    write_threshold_sweep(records, sweep_path)
    write_speed_quality_scatter(records, scatter_path)

    print(
        f"Wrote {long_path.name}, {f1_path.name}, {f1_thr_path.name}, "
        f"{heatmap_path.name}, {hist_path.name}, {sweep_path.name}, "
        f"{scatter_path.name} -> {RESULTS_DIR}"
    )

    print_markdown_table(
        records,
        value_fn=lambda r: f"{r['F1']:.3f} @ {r['threshold']:.2f}",
        title="Best F1 per (model, reference text)  —  cell = F1 @ threshold",
    )
    print_markdown_table(
        records,
        value_fn=lambda r: f"{r['F1']:.3f}",
        title="Best F1 only (clean version)",
    )


if __name__ == "__main__":
    main()
