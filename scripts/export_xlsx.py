"""Export processed JSON filings to XLSX summary."""

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter


"""
standalone script to export a XLSX file with context windows for manual classification
from the processed JSON files. run with python scripts/export_xlsx.py
only takes into account 1 context window for each filling. derives company name and ticker
from the "display_names" field in the filing metadata, which has the format:
"COMPANY NAME  (TICKER)  (CIK ...)"
"""

PROCESSED_BASE = Path(__file__).resolve().parent.parent / "data" / "json_processed"
VERSION = "version_1"
DATA_DIR = PROCESSED_BASE / VERSION / "files"
OUTPUT_PATH = Path("filings_summary.xlsx")


# Requires "COMPANY NAME  (TICKER)  (CIK ...)" — two parenthesized groups
DISPLAY_NAME_RE = re.compile(r"^(.+?)\s*\(([^)]+)\)\s*\(CIK\s+[^)]+\)")


def parse_display_name(name: str):
    """Return (company_name, ticker) or None if format doesn't match."""
    m = DISPLAY_NAME_RE.match(name)
    if not m:
        return None
    company = m.group(1).strip()
    ticker = m.group(2).strip()
    return company, ticker


def collect_rows(only_relevant: bool = True):
    from nps_crawling.db.db_adapter import DbAdapter
    db = DbAdapter()
    db_rows = db.get_all_filings()
    
    rows = []
    for row in db_rows:
        if only_relevant and not row.get("nps_relevant"):
            continue
            
        display_names = row.get("display_names") or []
        display_name = display_names[0] if display_names else ""
        parsed = parse_display_name(display_name)
        if parsed:
            company_name, ticker = parsed
        else:
            company_name = display_name
            tickers = row.get("ticker") or []
            ticker = tickers[0] if tickers else ""
            
        ciks = row.get("ciks") or []
        cik = ciks[0] if ciks else ""
        
        filing_type = row.get("file_type") or ""
        filing_date = row.get("file_date") or ""
        url = row.get("url") or ""
        
        snippet = ""
        path_to_preprocessed = row.get("path_to_preprocessed")
        if path_to_preprocessed:
            p = Path(path_to_preprocessed)
            if not p.is_absolute():
                p = Path(__file__).resolve().parent.parent / p
            if p.exists():
                try:
                    with open(p, encoding="utf-8") as f:
                        records = json.load(f)
                        if records:
                            record = records[0]
                            contexts = record.get("context", [])
                            if contexts:
                                snippet = contexts[0].get("context", "")
                                snippet = ILLEGAL_CHARACTERS_RE.sub("", snippet)
                except Exception:
                    pass
                    
        rows.append([
            company_name,
            ticker,
            cik,
            filing_type,
            filing_date,
            url,
            snippet,
        ])

    return rows


def main(only_relevant: bool = True):
    rows = collect_rows(only_relevant=only_relevant)
    if not rows:
        print("No valid records found.")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"

    headers = [
        "company_name",
        "ticker",
        "cik",
        "filing_type",
        "filing_date",
        "filing_url",
        "snippet_text_short",
    ]
    ws.append(headers)

    for row in rows:
        ws.append(row)

    # Auto-width (capped at 60)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(OUTPUT_PATH)
    print(f"Saved {len(rows)} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main(only_relevant=True)
