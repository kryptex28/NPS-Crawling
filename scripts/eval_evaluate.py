"""Compare manual labels against scored context windows.

Joins ``data/evaluation/labeling.xlsx`` (your manual labels, keyed by
``context_id``) with one or more ``data/evaluation/scores_*.jsonl`` files
(produced by ``eval_score_contexts.py``) and prints precision / recall / F1
across a sweep of thresholds for each scored model.

Usage:
    python scripts/eval_evaluate.py
    python scripts/eval_evaluate.py --scores scores_all-MiniLM-L6-v2.jsonl
"""

import argparse
import json
import sys
from pathlib import Path

# Allow running as a plain script without installing the package
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from nps_crawling.config import Config


EVAL_DIR = Config.DATA_PATH / "evaluation"
LABELING_XLSX = EVAL_DIR / "labeling.xlsx"


def load_labels():
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl is not installed. Run:  pip install openpyxl")
        sys.exit(1)

    if not LABELING_XLSX.exists():
        print(f"Missing {LABELING_XLSX}. Run eval_export_for_labeling.py first.")
        sys.exit(1)

    wb = load_workbook(LABELING_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    cid_col = headers.index("context_id")
    label_col = headers.index("label")

    labels = {}
    for r in rows:
        cid = r[cid_col]
        label = r[label_col]
        if cid is None or label is None or str(label).strip() == "":
            continue
        try:
            labels[str(cid)] = int(label)
        except (TypeError, ValueError):
            continue
    return labels


def load_scores(path: Path):
    out = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            out[obj["context_id"]] = obj["similarity_score"]
    return out


def metrics(labels, scores, threshold):
    tp = fp = fn = tn = 0
    for cid, label in labels.items():
        if cid not in scores:
            continue
        pred = 1 if scores[cid] >= threshold else 0
        if pred == 1 and label == 1:
            tp += 1
        elif pred == 1 and label == 0:
            fp += 1
        elif pred == 0 and label == 1:
            fn += 1
        else:
            tn += 1
    prec = tp / (tp + fp) if (tp + fp) else 0.0
    rec = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) else 0.0
    return tp, fp, fn, tn, prec, rec, f1


def evaluate_one(scores_path: Path, labels: dict, thresholds):
    scores = load_scores(scores_path)
    matched = sum(1 for cid in labels if cid in scores)

    print(f"\n=== {scores_path.name} ===")
    print(f"Labels: {len(labels)}  |  Scored & labeled: {matched}  |  "
          f"Missing scores for labeled rows: {len(labels) - matched}")
    if matched == 0:
        print("  No overlap between labels and scores — check context_ids.")
        return

    print(f"\n{'thr':>6} {'TP':>5} {'FP':>5} {'FN':>5} {'TN':>5} "
          f"{'P':>7} {'R':>7} {'F1':>7}")
    best = (None, -1.0)
    for t in thresholds:
        tp, fp, fn, tn, p, r, f1 = metrics(labels, scores, t)
        print(f"{t:>6.2f} {tp:>5} {fp:>5} {fn:>5} {tn:>5} "
              f"{p:>7.3f} {r:>7.3f} {f1:>7.3f}")
        if f1 > best[1]:
            best = (t, f1)
    print(f"\n  Best F1 at threshold {best[0]:.2f}: F1={best[1]:.3f}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--scores", nargs="*",
        help="Specific scores_*.jsonl file(s) (relative to data/evaluation/). "
             "Default: all scores_*.jsonl found.",
    )
    parser.add_argument(
        "--min", type=float, default=0.00,
        help="Lowest threshold in sweep (default: 0.00)",
    )
    parser.add_argument(
        "--max", type=float, default=0.50,
        help="Highest threshold in sweep (default: 0.50)",
    )
    parser.add_argument(
        "--step", type=float, default=0.02,
        help="Threshold step (default: 0.02)",
    )
    args = parser.parse_args()

    labels = load_labels()
    if not labels:
        print("No labels found in the xlsx. Fill in the 'label' column first.")
        return
    pos = sum(1 for v in labels.values() if v == 1)
    print(f"Loaded {len(labels)} labels  ({pos} positive, {len(labels) - pos} negative)")

    if args.scores:
        score_files = [EVAL_DIR / name for name in args.scores]
    else:
        score_files = sorted(EVAL_DIR.glob("scores_*.jsonl"))
    if not score_files:
        print(f"No scores_*.jsonl in {EVAL_DIR}. Run eval_score_contexts.py first.")
        return

    thresholds = []
    t = args.min
    while t <= args.max + 1e-9:
        thresholds.append(round(t, 4))
        t += args.step

    for sp in score_files:
        if not sp.exists():
            print(f"Missing {sp}, skipping.")
            continue
        evaluate_one(sp, labels, thresholds)


if __name__ == "__main__":
    main()
