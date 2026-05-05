"""Export context windows from json_raw into an Excel file for manual labeling.

Runs ONLY cleaning + filtering (no similarity scoring) over every JSON file in
``data/json_raw/files/`` and writes:

    data/evaluation/contexts.jsonl   — frozen source of truth (one row per context)
    data/evaluation/labeling.xlsx    — same rows, with an empty ``label`` column

You manually fill the ``label`` column in Excel (1 = relevant, 0 = not relevant,
leave empty to skip). The ``context_id`` column ties labels back to scores
produced later by ``eval_score_contexts.py``.

Re-running this script regenerates both files. Don't regenerate after you've
started labeling unless you also re-do labels — IDs are stable as long as
cleaning/filtering logic and json_raw contents don't change.
"""

import json
import sys
from pathlib import Path

from tqdm import tqdm

# Allow running as a plain script without installing the package
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from nps_crawling.config import Config
from nps_crawling.preprocessing.cleaning import CleanTextPipeline
from nps_crawling.preprocessing.filtering import NpsMentionFilterPipeline


EVAL_DIR = Config.DATA_PATH / "evaluation"
CONTEXTS_JSONL = EVAL_DIR / "contexts.jsonl"
LABELING_XLSX = EVAL_DIR / "labeling.xlsx"


def build_context_id(filename_stem: str, record_idx: int, context_idx: int) -> str:
    return f"{filename_stem}__r{record_idx}__c{context_idx}"


def extract_rows():
    cleaner = CleanTextPipeline()
    filt = NpsMentionFilterPipeline()
    raw_dir = Config.RAW_JSON_PATH_CRAWLER / "files"
    files = sorted(raw_dir.glob("*.json"))
    if not files:
        print(f"No files in {raw_dir}")
        return []

    rows = []
    for path in tqdm(files, desc="Extracting context windows", unit="file"):
        try:
            with open(path, "r", encoding="utf-8") as f:
                records = json.load(f)
        except json.JSONDecodeError:
            continue
        if not records:
            continue

        records = cleaner.cleaning_workflow(records)
        records = filt.filtering_workflow(records)

        stem = path.stem
        for r_idx, record in enumerate(records):
            meta = record.get("metadata", {}) or {}
            filing = meta.get("filing", {}) or {}
            for c_idx, ctx in enumerate(record.get("context", [])):
                rows.append({
                    "context_id": build_context_id(stem, r_idx, c_idx),
                    "source_file": stem,
                    "filing_id": filing.get("id"),
                    "company": filing.get("company") or filing.get("name"),
                    "form": filing.get("form"),
                    "filing_date": filing.get("filing_date") or filing.get("date"),
                    "matched_phrase": ctx.get("matched_phrase"),
                    "context": ctx.get("context", ""),
                })
    return rows


def write_jsonl(rows, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows, path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("openpyxl is not installed. Run:  pip install openpyxl")
        print(f"Skipping XLSX. JSONL still written to: {CONTEXTS_JSONL}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "labeling"

    headers = [
        "context_id", "label", "notes",
        "source_file", "filing_id", "company", "form", "filing_date",
        "matched_phrase", "context",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([
            row["context_id"], "", "",
            row["source_file"], row["filing_id"], row["company"],
            row["form"], row["filing_date"],
            row["matched_phrase"], row["context"],
        ])

    widths = {
        "A": 38, "B": 8, "C": 24,
        "D": 28, "E": 14, "F": 28, "G": 8, "H": 14,
        "I": 22, "J": 120,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    for row_cells in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row_cells:
            cell.alignment = wrap

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    rows = extract_rows()
    if not rows:
        print("No context windows extracted — nothing to write.")
        return
    write_jsonl(rows, CONTEXTS_JSONL)
    write_xlsx(rows, LABELING_XLSX)
    print(f"\nExtracted {len(rows)} context windows.")
    print(f"  JSONL : {CONTEXTS_JSONL}")
    print(f"  XLSX  : {LABELING_XLSX}")
    print("\nNext: open the XLSX, fill the 'label' column (1=relevant, 0=not).")


if __name__ == "__main__":
    main()
