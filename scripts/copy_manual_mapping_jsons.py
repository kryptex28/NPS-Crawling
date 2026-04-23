"""Copy raw JSON files whose URL appears in manual_mapping.xlsx to scripts/manual_mapping/."""

import json
import os
import shutil

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "manual_mapping.xlsx")
JSON_RAW_DIR = os.path.join(SCRIPT_DIR, "..", "data", "json_raw", "files")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "manual_mapping")


def get_excel_urls(excel_path: str) -> set[str]:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    idx = headers.index("filing_url")
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx]:
            urls.add(row[idx].strip())
    return urls


def copy_matching_jsons(excel_urls: set[str], json_dir: str, out_dir: str) -> int:
    os.makedirs(out_dir, exist_ok=True)
    matched = 0
    for fname in os.listdir(json_dir):
        if not fname.endswith(".json"):
            continue
        fpath = os.path.join(json_dir, fname)
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)
        for item in data:
            url = item.get("url", "")
            if url and url.strip() in excel_urls:
                shutil.copy2(fpath, os.path.join(out_dir, fname))
                matched += 1
                break
    return matched


def main():
    excel_urls = get_excel_urls(EXCEL_PATH)
    print(f"URLs in Excel: {len(excel_urls)}")

    matched = copy_matching_jsons(excel_urls, JSON_RAW_DIR, OUTPUT_DIR)
    print(f"Matched and copied: {matched} files")
    print(f"Output directory: {os.path.abspath(OUTPUT_DIR)}")


if __name__ == "__main__":
    main()
